from operator import pos
from docutils.nodes import line
from pyvi.pyvi import ViTokenizer
import openpyxl
from nltk import word_tokenize,ngrams
import re

def load_data(filePath):
    data = []
    with open(filePath,"r") as file:
        lines = file.readlines()
        for line in lines:
            if line!="":
                arr = line.split("\t")
                data.append(arr[7].strip())
    return data


def load_dic_sentiment(filePath):
    pos_dic = []
    neg_dic = []
    total_dic = []
    twin_dic = []
    wb = openpyxl.load_workbook(filePath, read_only=True)
    sheet_names = wb.sheetnames
    sheet = wb[sheet_names[0]]
    for row in sheet.iter_rows(row_offset=1) :
        if(row[2].value == 1) :
            pos_dic.append(row[1].value)
        if(row[3].value == 1) :
            neg_dic.append(row[1].value)
        if(row[2].value == 1 and row[3].value == 1) :
            twin_dic.append(row[1].value)
    total_dic.extend(neg_dic)
    total_dic.extend(pos_dic)
    print("len neg_dic : {}".format(len(neg_dic)))
    print("len pos_dis : {}".format(len(pos_dic)))
    print("len total_dic : {}".format(len(total_dic)))
    print("len twin_dic : {}".format(len(twin_dic)))
    wb.close()
    return total_dic,pos_dic,neg_dic


def format_data(raw_data):
    data = []
    pattern = r"^[^-_@\.\$%\(\)\[\]](.*)[^-_@\.\$%\(\)\[\]\\]$"
    for line in raw_data :
        line = re.sub("_", " ", line)
        line = word_tokenize(line.lower())
        new_line = []
        for word in line:
            if re.match(pattern,word) :
                new_line.append(word)
        data.append(" ".join(new_line))
        # data.append(new_line)
    return data


def compute(data,total_dic, pos_dic, neg_dic):
    onegram_data = transform(data,1)
    twogram_data = transform(data,2)
    threegram_data = transform(data,3)
    fourgram_data = transform(data,4)
    point_data_total = []
    point_data1 = compute_perdata(onegram_data,total_dic,pos_dic,neg_dic)
    point_data2 = compute_perdata(twogram_data,total_dic,pos_dic,neg_dic)
    point_data3 = compute_perdata(threegram_data,total_dic,pos_dic,neg_dic)
    point_data4 = compute_perdata(fourgram_data,total_dic,pos_dic,neg_dic)
    for i in range(len(point_data1)) :
        tup1 = point_data1[i]
        tup2 = point_data2[i]
        tup3 = point_data3[i]
        tup4 = point_data4[i]
        point_data_total.append((tup1[0] + tup2[0] + tup3[0] +tup4[0],
                                tup1[1] + tup2[1] + tup3[1] + tup4[1]))
    return point_data_total



def compute_perdata(data,total_dic, pos_dic, neg_dic):
    point_data = []
    for line in data :
        point_pos = 0
        point_neg = 0
        # print(line)
        # print(pos_dic)
        for word in line :
            if(word in neg_dic) :
                point_neg +=1
            elif(word in pos_dic) :
                point_pos +=1
        point_data.append((point_pos,point_neg))
    return point_data


def transform(raw_data,n):
    data = []
    for line in raw_data :
        new_line = []
        ngram = ngrams(line.split(), n)
        for grams in ngram:
            new_line.append(" ".join(grams))
        data.append(new_line)
    return data


if __name__ == "__main__" :
    filePath = "data/ButViet.txt"
    raw_data = load_data(filePath)
    data = format_data(raw_data)
    # for line in data:
    #     print(line)
    print("len data :  {}".format(len(data)))
    file_path_sentiment = "/home/captainlt/Desktop/VnEmoLex.xlsx"
    total_dic, pos_dic, neg_dic = load_dic_sentiment(file_path_sentiment)
    point_data_total = compute(data,total_dic, pos_dic, neg_dic)
    count = 0
    for i in range(len(data)) :
        point = point_data_total[i]
        if point[0] < point[1] :
            print(data[i])
            print(point)
            count +=1
    print(count)
        # print(data[i])
        # print(point_data_total[i])
